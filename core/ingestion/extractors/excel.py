"""Excel content extractor using openpyxl."""

from __future__ import annotations

import io
import logging
from pathlib import Path
from typing import Union

from openpyxl import load_workbook

from core.exceptions import ExtractionError
from core.ingestion.extractors.base import BaseExtractor, ExtractedDocument

logger = logging.getLogger(__name__)


class ExcelExtractor(BaseExtractor):
    """Extracts text content from Excel (.xlsx) spreadsheets."""

    supported_extensions = [".xlsx", ".xls"]

    async def extract(self, source: Union[str, Path, bytes]) -> ExtractedDocument:
        """Extract text from an Excel file or byte stream.

        Each sheet is rendered as a Markdown-style table.

        Args:
            source: File path or raw Excel bytes.

        Returns:
            ExtractedDocument with tabular text.

        Raises:
            ExtractionError: If the workbook cannot be read.
        """
        try:
            if isinstance(source, bytes):
                wb = load_workbook(io.BytesIO(source), read_only=True, data_only=True)
                file_name = "uploaded.xlsx"
            else:
                path = Path(source)
                if not path.exists():
                    raise ExtractionError(f"Excel file not found: {path}")
                wb = load_workbook(str(path), read_only=True, data_only=True)
                file_name = path.name

            sections: list[str] = []
            total_rows = 0

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows: list[str] = []
                for row in ws.iter_rows(values_only=True):
                    cell_values = [str(c) if c is not None else "" for c in row]
                    if any(v.strip() for v in cell_values):
                        rows.append(" | ".join(cell_values))
                if rows:
                    header = f"## Sheet: {sheet_name}\n"
                    sections.append(header + "\n".join(rows))
                    total_rows += len(rows)

            wb.close()

            content = "\n\n".join(sections)
            metadata = {
                "source_type": "excel",
                "file_name": file_name,
                "sheet_count": len(wb.sheetnames),
                "total_rows": total_rows,
            }

            logger.info("Extracted %d rows from %d sheets in %s", total_rows, len(wb.sheetnames), file_name)
            return ExtractedDocument(content=content, metadata=metadata, title=file_name)

        except ExtractionError:
            raise
        except Exception as exc:
            raise ExtractionError(f"Failed to extract Excel: {exc}") from exc
